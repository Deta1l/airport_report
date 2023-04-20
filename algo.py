from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook, open


#schedule_xlsx = open('data\Schedule_LT.xlsx', read_only=True) 

mct_gruz = 'data\MCT_груз.xlsx'
schedule_xlsx1  = 'data\Schedule_LT.xlsx'

wb1 = load_workbook(schedule_xlsx1)
mct = load_workbook(mct_gruz)

answer = {}
city = []
route = []

ws1 = wb1.active
m_row = ws1.max_row

mct_ws = mct.active
mct_row = mct_ws.max_row

for i in range(2, m_row + 1):
    cell_obj = ws1.cell(row = i, column = 2)
    if cell_obj.value not in city:
        city.append(cell_obj.value)

        for a in range(2, m_row + 1):
            cell_obj1 = ws1.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws1.cell(row = a, column = 4)
                answer[cell_obj.value]=[g1.value]
                break
        
        for a in range(2, m_row + 1):
            cell_obj1 = ws1.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws1.cell(row = a, column = 4)
                answer[cell_obj.value].append(g1.value)


def dfs_paths(graph, n, start, goal, path=[], count=0):
    path = path + [start]

    if start == goal and len(path) <= n+2:
        route.append(path)
        return
    
    for next_node in graph[start]:
        if next_node not in path:
            #проверка на время + время стоянки
            dfs_paths(graph, n, next_node, goal, path, count+1)


graph = {'a':['b'],
         'b':['a', 'c', 'c', 'c'],
         'c':['b', 'b', 'b']
    
}

start = 'OVB'
end = 'EVN'
n = 2

dfs_paths(answer, n , start, end, [], 0)
print(route)

print("finish")

'''

from datetime import datetime as dt 

for i in range(2, m_row + 1):
    cell_obj = ws.cell(row = i, column = 3)
    cell_obj1 = ws.cell(row = i, column = 5)
    time_1 = dt.strptime(str(cell_obj.value),"%H:%M:%S")
    time_2 = dt.strptime(str(cell_obj1.value),"%H:%M:%S")
    time_interval = time_2 - time_1
    #print(time_interval)
'''

#--------------------------------вывод


wb1.close()