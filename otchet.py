from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook, load_workbook, open
from datetime import datetime

wb = Workbook()

schedule_xlsx = open('data\Schedule_LT.xlsx', read_only=True) 
print(schedule_xlsx)

ws = wb.active

ws.merge_cells('A1:B4')
start = ws['A1']
start.value = "OnD"
start.alignment = Alignment(horizontal="center", vertical="center")


#от до
ws.merge_cells('A5:B9')
ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
#ws['A5'] = input("введине направление ")

#стыковка
ws.merge_cells('C5:C9')


ws.merge_cells('C2:J2')
s7 = ws['C2']
s7.value = "S7"
s7.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells('C3:C4')
ws['C3'] = 'Город стыковки'
ws.column_dimensions['C'].width = 15
ws.merge_cells('D3:D4')
ws['D3'] = '№ рейсов'
ws.column_dimensions['D'].width = 15
ws.merge_cells('E3:E4')
ws['E3'] = 'период'
ws.column_dimensions['E'].width = 15
ws.merge_cells('F3:F4')
ws['F3'] = 'дни недели'
ws.column_dimensions['F'].width = 15
ws.merge_cells('G3:G4')
ws['G3'] = 'вылет'
ws.column_dimensions['G'].width = 15
ws.merge_cells('H3:H4')
ws['H3'] = 'прилет'
ws.column_dimensions['H'].width = 15
ws.merge_cells('I3:I4')
ws['I3'] = 'время стык'
ws.column_dimensions['I'].width = 15
ws.merge_cells('J3:J4')
ws['J3'] = 'время полета'
ws.column_dimensions['J'].width = 15


#schedule_xlsx = open('data\Schedule_LT.xlsx', read_only=True) 

mct_gruz = 'data\MCT_груз.xlsx'
schedule_xlsx1  = 'data\Schedule_LT.xlsx'

wb1 = load_workbook(schedule_xlsx1)
mct = load_workbook(mct_gruz)

answer = {}
city = []
route = []

ws1 = wb1.active
m_row = ws1.max_row

mct_ws = mct.active
mct_row = mct_ws.max_row

for i in range(2, m_row + 1):
    cell_obj = ws1.cell(row = i, column = 2)
    if cell_obj.value not in city:
        city.append(cell_obj.value)

        for a in range(2, m_row + 1):
            cell_obj1 = ws1.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws1.cell(row = a, column = 4)
                answer[cell_obj.value]=[g1.value]
                break
        
        for a in range(2, m_row + 1):
            cell_obj1 = ws1.cell(row = a, column = 2)
            if cell_obj1.value == cell_obj.value:
                g1 = ws1.cell(row = a, column = 4)
                answer[cell_obj.value].append(g1.value)


def dfs_paths(graph, n, start, goal, path=[], count=0):
    path = path + [start]

    if start == goal and len(path) <= n+2:
        route.append(path)
        return
    
    for next_node in graph[start]:
        if next_node not in path:
            #проверка на время + время стоянки
            dfs_paths(graph, n, next_node, goal, path, count+1)


start = 'OVB'
end = 'EVN'
n = 2

dfs_paths(answer, n , start, end, [], 0)
print(route)

print("finish")


#--------------------------------вывод

lines_number = 5

for one_route in route:
 
    if len(one_route) == 2:
        ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C5'] = 'Direct Flight'
        start_airport = one_route[0]
        end_airport = one_route[1]
        for i in range(2, m_row + 1):
            cell_obj = ws1.cell(row = i, column = 2)
            cell_obj1 = ws1.cell(row = i, column = 4)
            if (start_airport == cell_obj.value) and (end_airport == cell_obj1.value):
                flight_number = ws1.cell(row = i, column = 1) #ВЫВОД
                ws.cell(row = lines_number, column = 4, value = flight_number.value)
                date1 = ws1.cell(row = i, column = 6)
                date1 = str(date1.value)
                flite_dates, flight_reg = date1.split()  #2ВЫВОД
                ws.cell(row = lines_number, column = 5, value = flite_dates)
                ws.cell(row = lines_number, column = 6, value = flight_reg)
            
                dep = ws1.cell(row = i, column = 3) #ВЫВОД
                arr = ws1.cell(row = i, column = 5) #ВЫВОД

                print(arr.value - dep.value)

                ws.cell(row = lines_number, column = 7, value = dep.value)
                ws.cell(row = lines_number, column = 8, value = arr.value)
                ws.cell(row = lines_number, column = 9, value = 0)
                lines_number+=1
                #длительность стыковыки 0, посчитать время полета

         





print("finish")

wb1.close()


wb.save('otchet.xlsx')
wb.close()
print("отчет готов!")